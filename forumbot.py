from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import openpyxl

chrome_options = Options()
chrome_options.add_argument("--headless")
chrome_options.add_argument("start-maximized")
chrome_options.add_argument("--disable-blink-features")
chrome_options.add_argument("--disable-blink-features=AutomationControlled")




username = "****"  #forum kullanıcı adı
password = "****" #forum şifresi




options = Options()
options.binary_location = r'C:\Program Files\Mozilla Firefox\firefox.exe' # firefox taraycısının tam bağlantı yolu olan exe yolunu ekliyorsunuz.
driver = webdriver.Firefox(executable_path=r'C:\Users\***\***\****\geckodriver.exe', options=options)  #geckodriver.exenin tam bağlantı yolunu atıyorsunuz.
#driver = webdriver.Chrome(ChromeDriverManager().install())
wait = WebDriverWait(driver, 15)
driver.get("https://forum.sorrymother.video/login")
wait.until(EC.element_to_be_clickable((By.NAME, "login"))).send_keys(username)
wait.until(EC.element_to_be_clickable((By.NAME, "password"))).send_keys(password)
wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@class = 'button--primary button button--icon button--icon--login'] "))).click() 
print("Logged in")



path = "frmicin.xlsx" # içeriklerin olduğu exceli yazıyorsunuz.

# workbook object is created
wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
for i in range(1, m_row + 1):
    title=sheet_obj.cell(row = i, column = 1).value
    link=sheet_obj.cell(row = i, column = 3).value
    with open("PaylasilanPostlar.txt", "r") as file3:  #Paylasilansublar.txt elle oluştur.
        Postlinks = file3.read()
        Postlinks = Postlinks.split("\n")
        Postlinks = list(filter(None, Postlinks))
        if link not in Postlinks :
            print("\n\n YENİ İÇERİK BULUNDU -> ",link,"\n")
            with open ("PaylasilanPostlar.txt", "a") as f:  
                f.write(link + "\n")
            if (50<len(title)):
                title=title[:45]
            link ="[REPLY]"+ link + "[/REPLY]" + "\n\n [B]JOIN FOR MORE[/B] : Telegram channel : https://t.me/+ikgWFnONfzsxZGY0\n\n "  # Buraya telegram linkiniz gelecek.  
            print(title)
            
            Threads=["https://forum.sorrymother.video/forums/onlyfans.5/"
                    ,"https://forum.sorrymother.video/forums/requests.4/"
                    ,"https://forum.sorrymother.video/forums/patreon.6/"
                    ,"https://forum.sorrymother.video/forums/youtubers.7/"
                    ,"https://forum.sorrymother.video/forums/twitch.8/"
                    ,"https://forum.sorrymother.video/forums/snapchat.9/"
                    ,"https://forum.sorrymother.video/forums/instagram.10/"
                    ,"https://forum.sorrymother.video/forums/manyvids.11/"
                    ,"https://forum.sorrymother.video/forums/xxx-full-site-rips.12/"
                    ,"https://forum.sorrymother.video/forums/cams.14/"
                    ,"https://forum.sorrymother.video/forums/twitter.15/"
                    ,"https://forum.sorrymother.video/forums/reddit.16/"
                    ,"https://forum.sorrymother.video/forums/hentai.17/"
                    ,"https://forum.sorrymother.video/forums/celeb.18/"
                    ,"https://forum.sorrymother.video/forums/ts-trap-shemale.23/"
                    ,"https://forum.sorrymother.video/forums/tiktok.53/"
                    ,"https://forum.sorrymother.video/forums/amateur.63/"
                ]
            for Thread in Threads:
                
                driver.get(Thread+"post-thread")
                wait.until(EC.element_to_be_clickable((By.NAME, "title"))).send_keys(title)
                time.sleep(2)
                wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class = 'fr-element fr-view fr-element-scroll-visible'] "))).send_keys(link)
                time.sleep(2)
                wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@class ='select2-search__field']"))).send_keys("rare")
                time.sleep(7)
                
                wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@class ='select2-search__field']"))).send_keys(Keys.ENTER)
                time.sleep(4)
#                 wait.until(EC.element_to_be_clickable((By.XPATH, "//li[@class ='select2-selection__choice']"))).click()

                wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@class = 'button--primary button button--icon button--icon--write'] "))).send_keys(Keys.ENTER)
                time.sleep(3)
                print("POST SENT")
                time.sleep(5)
                driver_len = len(driver.window_handles) #fetching the Number of Opened tabs
                print("Length of Driver = ", driver_len)
                if driver_len > 1: # Will execute if more than 1 tabs found.
                    for i in range(driver_len - 1, 0, -1):
                        driver.switch_to.window(driver.window_handles[i]) #will close the last tab first.
                        driver.close()
                        print("Closed Tab No. ", i)
                    driver.switch_to.window(driver.window_handles[0]) # Switching the driver focus to First tab.
                else:
                    print("Found only Single tab.")
                time.sleep(260)
            time.sleep(60)
# //*[@id="top"]/div[4]/div/div[2]/div/div/div/div[2]/div[2]/div/form/div[2]/div/dl[2]/dd/span/span[1]/span            
